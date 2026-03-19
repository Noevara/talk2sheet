from __future__ import annotations

import csv
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .loader_common import normalize_sheet_index, path_cache_key, read_excel_frame


def _normalize_data_start_row(header_plan: dict[str, Any] | None) -> int:
    if not isinstance(header_plan, dict):
        return 2
    return max(1, int(header_plan.get("data_start_row_1based") or 2))


def _has_row_content(row: tuple[Any, ...] | list[Any]) -> bool:
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str):
            if cell.strip():
                return True
            continue
        return True
    return False


@lru_cache(maxsize=128)
def _count_sheet_rows_cached(cache_key: tuple[str, int, int], *, sheet_index: int, data_start_row_1based: int) -> int:
    path = Path(cache_key[0])
    suffix = path.suffix.lower()
    if suffix == ".csv":
        row_count = 0
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for line_number, row in enumerate(reader, start=1):
                if line_number < data_start_row_1based:
                    continue
                if _has_row_content(row):
                    row_count += 1
        return row_count

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            idx = normalize_sheet_index(sheet_index, len(workbook.sheetnames))
            worksheet = workbook[workbook.sheetnames[idx]]
            row_count = 0
            for row in worksheet.iter_rows(min_row=max(1, data_start_row_1based), values_only=True):
                if _has_row_content(row):
                    row_count += 1
            return row_count
        finally:
            workbook.close()

    if suffix == ".xls":
        raw, _sheet_name = read_excel_frame(path, normalize_sheet_index(sheet_index), header_row=None, nrows=None)
        start = max(0, data_start_row_1based - 1)
        data = raw.iloc[start:]
        return int(data.dropna(axis=0, how="all").shape[0])

    raise RuntimeError(f"Unsupported spreadsheet format: {suffix}")


def count_sheet_rows(path: Path, *, sheet_index: int, header_plan: dict[str, Any] | None = None) -> int:
    return _count_sheet_rows_cached(
        path_cache_key(path),
        sheet_index=max(1, int(sheet_index or 1)),
        data_start_row_1based=_normalize_data_start_row(header_plan),
    )
